from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash, Response, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask import Blueprint
import re
import json
from flask_login import current_user, login_required
from sqlalchemy.exc import IntegrityError
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split    
from sklearn.ensemble import IsolationForest, RandomForestClassifier
import re
import matplotlib.pyplot as plt
import seaborn as sns
from flask import send_from_directory
from utils.history_utils import get_browser_history_paths, extract_history_from_browser
from io import BytesIO
from werkzeug.utils import secure_filename
import platform
import psutil
import getpass
import subprocess
import os
import io
import time
import webbrowser
from browser_history import get_history
from collections import defaultdict, deque
import threading
from datetime import datetime, timedelta
from forms import RegistrationForm, LaptopRegistrationForm
import sqlite3
import glob
import winreg 
import time
import hashlib
import re
from urllib.request import Request, urlopen
from urllib.parse import urlparse
import urllib.parse
import hashlib
from collections import defaultdict
from pathlib import Path
from flask import make_response
from openpyxl import Workbook, load_workbook
from io import BytesIO
from functools import lru_cache
from typing import Dict, List
from sqlalchemy import func
from models import AuthLog
from utils.firebase_db import FirebaseDB


user_bp = Blueprint('user', __name__, url_prefix='/user')

# Define the filter function
def basename_filter(path):
    return os.path.basename(path)

# Register the filter for this blueprint
user_bp.add_app_template_filter(basename_filter, 'basename')


def get_system_serial():
    try:
        if platform.system() == "Windows":
            # For Windows (using powershell as wmic is deprecated/unavailable)
            result = subprocess.check_output('powershell.exe -Command "Get-CimInstance Win32_BIOS | Select-Object -ExpandProperty SerialNumber"', shell=True)
            return result.decode().strip()
        elif platform.system() == "Linux":
            # For Linux
            with open('/sys/class/dmi/id/product_serial', 'r') as f:
                return f.read().strip()
        elif platform.system() == "Darwin":
            # For macOS
            result = subprocess.check_output(['system_profiler', 'SPHardwareDataType'])
            for line in result.decode().split('\n'):
                if "Serial Number" in line:
                    return line.split(":")[1].strip()
    except:
        return "Not Available"
    return "Not Available"

@user_bp.route('/home', methods=['GET', 'POST'])
def home():
    df = get_ml_df()
    if current_user.role != 'Admin':
        df = df[df['Username'] == current_user.username]
    import platform, getpass, psutil, json, re

    # --- System info ---
    system_info = {
        'os_name': platform.system(),
        'os_version': platform.version(),
        'os_release': platform.release(),
        'machine_type': platform.machine(),
        'processor': platform.processor(),
        'hostname': platform.node(),
        'username': getpass.getuser(),
        'serial_number': get_system_serial(),
        'ram_total': round(psutil.virtual_memory().total / (1024**3)),
        'ram_used': round(psutil.virtual_memory().used / (1024**3)),
        'ram_percent': psutil.virtual_memory().percent
    }

    # --- Critical alerts count ---
    target_categories = ['anonymous', 'betting', 'hacking', 'money', 'trading', 'porn']
    critical_alerts_count = len(df[
        (df['risk_level'].isin(['Critical'])) &
        (df[[f'is_{cat}' for cat in target_categories if f'is_{cat}' in df.columns]].any(axis=1))
    ])

    # --- Incident Volume by category ---
    threat_volume_by_category = {}

    # 1. Calculate Main Category Volume
    for cat_col in [c for c in df.columns if c.startswith('is_')]:
        category = cat_col.replace('is_', '')
        # Sum up the actual visit counts for this threat category
        volume = df[df[cat_col] == 1]['Visit Count'].sum()
        if volume > 0:
            threat_volume_by_category[category.capitalize()] = int(volume)

    # 2. Detailed Patterns grouping
    threat_patterns = {
        'betting': r'\b(bet|betting|gamb|casino|poker|sportsbook|wager|odds|slot)\b',
        'porn': r'\b(porn|xxx|adult|sex|fuck|nude|nsfw|webcam|escort)\b',
        'social': r'\b(facebook|instagram|twitter|tiktok|social[\s-]?media|linkedin|whatsapp)\b',
        'video': r'\b(youtube|vimeo|dailymotion|stream|twitch|netflix)\b',
        'trading': r'\b(forex|crypto|bitcoin|trading|binance|coinbase|kraken|mt[45])\b',
        'hacking': r'\b(hack|crack|keygen|cheat|exploit|bypass|ddos|injection)\b',
        'money': r'\b(western[\s-]?union|money[\s-]?gram|paypal|venmo|cash[\s-]?app)\b',
        'cloud': r'\b(dropbox|google[\s-]?drive|mega(?:\.nz)?|onedrive|box(?:\.com)?)\b',
        'anonymous': r'\b(tor|vpn|proxy|anonymous|incognito|hide[\s-]?ip)\b',
        'shopping': r'\b(amazon|ebay|alibaba|etsy|shopify|walmart|target)\b',
        'job': r'\b(linkedin|indeed|monster|career|job|employment)\b',
        'forum': r'\b(reddit|forum|4chan|discord|telegram)\b'
    }

    for main_category, pattern in threat_patterns.items():
        platforms = re.findall(r'\b([a-zA-Z0-9\[\]\-]+)\b', pattern)
        platforms = [p for p in platforms if len(p) > 2 and p.lower() not in ['social', 'media'] and not p.startswith('(')]

        for platform_name in platforms:
            col_name = f'is_{platform_name.lower()}'
            if col_name in df.columns:
                volume = df[df[col_name] == 1]['Visit Count'].sum()
                if volume > 0:
                    label = platform_name.capitalize()
                    # Only append main category if it's different from platform
                    if main_category.lower() != platform_name.lower():
                        label = f"{main_category.capitalize()} ({label})"
                    
                    threat_volume_by_category[label] = int(volume)

    # --- Prepare chart data ---
    # Sort by volume descending to show biggest threats first
    sorted_threats = sorted(threat_volume_by_category.items(), key=lambda x: x[1], reverse=True)
    
    chart_categories = [x[0] for x in sorted_threats]
    chart_scores = [x[1] for x in sorted_threats]

    # --- Admin Features ---
    # Fetch all users from Firestore
    global_users = FirebaseDB.get_global_users()
    total_users = len(global_users)
    
    from collections import defaultdict
    roles_counts_map = defaultdict(int)
    for u in global_users:
        role = u.get('role') or 'User'
        roles_counts_map[role] += 1
        
    roles_labels = list(roles_counts_map.keys())
    roles_counts = list(roles_counts_map.values())
    
    # Fetch auth logs from Firestore
    auth_logs = FirebaseDB.get_all_auth_logs(limit=2000)
    total_auth_events = len(auth_logs)
    success_events = sum(1 for log in auth_logs if log.get('action') in ['login_success', 'register_success'])
    failed_events = sum(1 for log in auth_logs if log.get('action') in ['login_failed', 'register_failed'])
    
    auth_labels = ['Successful', 'Possible Intrusion']
    auth_counts = [success_events, failed_events]

    return render_template(
        'user/home.html',
        system_info=system_info,
        critical_alerts_count=critical_alerts_count,
        categories_json=json.dumps(chart_categories),
        scores_json=json.dumps(chart_scores),
        total_users=total_users,
        roles_labels=roles_labels,
        roles_counts=roles_counts,
        auth_labels=auth_labels,
        auth_counts=auth_counts,
        total_auth_events=total_auth_events
    )



@user_bp.route('/profile')
@login_required
def profile():
    return render_template('user/profile.html', user=current_user)

@user_bp.route('/mobile', methods=['GET', 'POST'])
def mobile():
    return render_template('user/mobile.html', user=current_user)

@user_bp.route('/add-computers', methods=['GET', 'POST'])
@login_required
def add_computers():
    from forms import LaptopRegistrationForm
    form = LaptopRegistrationForm()
    if form.validate_on_submit():
        from firebase_config import db_firestore
        if db_firestore is not None:
            try:
                # Add laptop to firestore 'laptops' collection
                laptop_data = {
                    'full_name': form.full_name.data,
                    'email': form.email.data,
                    'laptop_name': form.laptop_name.data,
                    'laptop_model': form.laptop_model.data,
                    'serial_number': form.serial_number.data,
                    'laptop_os': form.laptop_os.data,
                    'status': 'active'
                }
                db_firestore.collection('laptops').add(laptop_data)
                flash('Endpoint registered successfully!', 'success')
                return redirect(url_for('user.computer_management'))
            except Exception as e:
                flash(f'Registration Error: {str(e)}', 'error')
        else:
            flash('Database Connection Error', 'error')
            
    return render_template('user/add_computers.html', form=form)

@user_bp.route('/computer-management', methods=['GET'])
@login_required
def computer_management():
    # Fetch from Firestore 'laptops'
    from firebase_config import db_firestore
    computers = []
    if db_firestore is not None:
        try:
            docs = db_firestore.collection('laptops').stream()
            for doc in docs:
                data = doc.to_dict()
                class LaptopAdapter:
                    def __init__(self, doc_id, d):
                        self.id = doc_id
                        self.full_name = d.get('full_name', '')
                        self.email = d.get('email', '')
                        self.laptop_name = d.get('laptop_name', '')
                        self.laptop_model = d.get('laptop_model', '')
                        self.serial_number = d.get('serial_number', '')
                        self.laptop_os = d.get('laptop_os', '')
                        self.status = d.get('status', 'active')
                computers.append(LaptopAdapter(doc.id, data))
        except Exception as e:
            print(f"Error fetching laptops: {e}")
            
    total_computers = len(computers)
    return render_template('user/computer_management.html', 
                         computers=computers,
                         total_computers=total_computers)
    
@user_bp.route('/update-laptop-status/<string:laptop_id>', methods=['POST'])
@login_required
def update_laptop_status(laptop_id):
    data = request.get_json()
    if data['status'] not in ['active', 'frozen']:
        return jsonify({'success': False, 'error': 'Invalid status'}), 400
    
    from firebase_config import db_firestore
    if db_firestore is not None:
        try:
            db_firestore.collection('laptops').document(laptop_id).update({
                'status': data['status']
            })
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    return jsonify({'success': False, 'error': 'Database Connection Error'}), 500

@user_bp.route('/delete-laptop/<string:laptop_id>', methods=['DELETE'])
@login_required
def delete_laptop(laptop_id):
    from firebase_config import db_firestore
    if db_firestore is not None:
        try:
            db_firestore.collection('laptops').document(laptop_id).delete()
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    return jsonify({'success': False, 'error': 'Database Connection Error'}), 500


# Configure export directory and filename (add this at the top of your file)
EXPORT_DIR = os.path.join(os.getcwd(), 'browser_history_exports')
EXPORT_FILENAME = "browser_history.xlsx"
os.makedirs(EXPORT_DIR, exist_ok=True)  # Create directory if it doesn't exist

def get_user_session_windows(username):
    """Helper to find all login/logout session windows for a given username from Firestore"""
    from datetime import datetime
    
    global_users = FirebaseDB.get_global_users()
    target_user = None
    for gu in global_users:
        if gu.get('username') == username:
            target_user = gu
            break
            
    if not target_user:
        return []
        
    user_id = target_user.get('id') or target_user.get('user_id')
    user_email = target_user.get('email')
    
    all_logs = FirebaseDB.get_all_auth_logs(limit=2000)
    
    user_logs = []
    for log in all_logs:
        if (user_email and log.get('email') == user_email) or \
           (user_id and str(log.get('user_id')) == str(user_id)):
            user_logs.append(log)
            
    def get_log_time(l):
        t = l.get('timestamp')
        if isinstance(t, datetime):
            return t.replace(tzinfo=None) if t.tzinfo is not None else t
        elif isinstance(t, str):
            try:
                return datetime.strptime(t[:19], '%Y-%m-%d %H:%M:%S')
            except ValueError:
                pass
        return datetime.min
        
    user_logs.sort(key=get_log_time)
    
    session_windows = []
    for i, log in enumerate(user_logs):
        if log.get('action') == 'login_success':
            start_time = get_log_time(log)
            if start_time == datetime.min:
                continue
            end_time = datetime.utcnow()
            for next_log in user_logs[i+1:]:
                if next_log.get('action') in ['logout', 'login_success']:
                    t_next = get_log_time(next_log)
                    if t_next != datetime.min:
                        end_time = t_next
                        break
            session_windows.append((start_time, end_time))
    return session_windows

@user_bp.route('/visited-webs', methods=['GET', 'POST'])
@login_required
def visited_webs():
    all_users = []
    if current_user.is_authenticated and current_user.role == 'Admin':
        global_users_data = FirebaseDB.get_global_users()
        all_users = []
        for gu in global_users_data:
            if gu.get('role') == 'User' and gu.get('username'):
                all_users.append({
                    'username': gu['username'],
                    'email': gu.get('email', 'Global User')
                })
        all_users = sorted(all_users, key=lambda x: x['username'])
    
    username = (
        request.form.get('username')
        or request.args.get('username')
        or (current_user.username if hasattr(current_user, "username") and current_user.is_authenticated else "Guest")
    )

    histories = []

    if request.method == 'POST' and username and username != "Guest":
        global_users = FirebaseDB.get_global_users()
        user_id = None
        for gu in global_users:
            if gu.get('username') == username:
                user_id = gu.get('id') or gu.get('user_id')
                break
        if not user_id:
            user_id = username

        # Clear existing history if requested from Firestore
        if request.form.get('clear_existing') == 'on':
            try:
                from firebase_config import db_firestore
                if db_firestore is not None:
                    batch = db_firestore.batch()
                    docs = db_firestore.collection('browser_history').where('user_id', '==', user_id).stream()
                    for d in docs:
                        batch.delete(d.reference)
                    docs_str = db_firestore.collection('browser_history').where('user_id', '==', str(user_id)).stream()
                    for d in docs_str:
                        batch.delete(d.reference)
                    docs_uname = db_firestore.collection('browser_history').where('username', '==', username).stream()
                    for d in docs_uname:
                        batch.delete(d.reference)
                    batch.commit()
                flash('Existing history cleared from Firestore', 'info')
            except Exception as e:
                print(f"Error clearing history: {e}")

        # Get this user's login sessions to ensure we only attribute history they actually visited
        session_windows = get_user_session_windows(username)
        
        # Extract history from all browsers
        browser_paths = get_browser_history_paths()
        total_items = 0
        history_data = []  # Store data for Excel export

        for browser, path in browser_paths:
            # Extract ALL history first
            raw_items = extract_history_from_browser(browser, path, user_id)
            
            # FILTER: Only keep items that happened during an active session for this user
            items = []
            for item_dict in raw_items:
                visit_time = item_dict['last_visit_time']
                is_valid_session = False
                for start, end in session_windows:
                    if start <= visit_time <= end:
                        is_valid_session = True
                        break
                if is_valid_session:
                    items.append(item_dict)
            
            total_items += len(items)
            
            for item in items:
                history_data.append({
                    'username': username,
                    'Browser': browser,
                    'Title': item['title'],
                    'URL': item['url'],
                    'Visit Count': item['visit_count'],
                    'Last Visit': item['last_visit_time'].strftime('%Y-%m-%d %H:%M:%S')
                })

        if history_data:
            # Sync to Firebase in background
            import threading
            threading.Thread(
                target=FirebaseDB.save_browser_history,
                args=(user_id, history_data),
                daemon=True
            ).start()

        flash(f'Successfully extracted {total_items} history items', 'success')
        
        # CRITICAL: Clear the ML cache so the next visit to Anomalies/Threats uses this new data
        global ml_df
        ml_df = None

        if history_data:
            try:
                excel_path = os.path.join(EXPORT_DIR, EXPORT_FILENAME)
                df_export = pd.DataFrame(history_data)
                update_ml_pipeline()  # Update ML model with new data
                df_export.to_excel(excel_path, index=False)
                flash('Browser history automatically updated in browser_history.xlsx', 'info')
            except Exception as e:
                flash(f'Error updating Excel file: {str(e)}', 'error')

    # Fetch history for display if user exists (handles both GET and POST)
    if username and username != "Guest":
        global_history = FirebaseDB.get_browser_history_by_username(username)
        if global_history:
            from datetime import datetime
            class FirebaseHistoryAdapter:
                def __init__(self, data):
                    self.browser = data.get('Browser') or data.get('browser') or ''
                    self.title = data.get('Title') or data.get('title') or ''
                    self.url = data.get('URL') or data.get('url') or ''
                    self.visit_count = data.get('Visit Count') or data.get('visit_count') or 1
                    
                    last_visit = data.get('Last Visit') or data.get('last_visit') or data.get('last_visit_time')
                    if isinstance(last_visit, str):
                        try:
                            self.last_visit_time = datetime.strptime(last_visit, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            self.last_visit_time = datetime.utcnow()
                    elif last_visit:
                        if isinstance(last_visit, datetime) and last_visit.tzinfo is not None:
                            self.last_visit_time = last_visit.replace(tzinfo=None)
                        else:
                            self.last_visit_time = last_visit
                    else:
                        self.last_visit_time = datetime.utcnow()
                            
            histories = [FirebaseHistoryAdapter(h) for h in global_history]
        else:
            histories = []
    else:
        histories = []

    return render_template('user/visited_webs.html',
                           histories=histories,
                           username=username,
                           all_users=all_users)

    return render_template('user/visited_webs.html',
                           histories=histories,
                           username=username,
                           all_users=all_users)

@user_bp.route('/exports/browser_history')
def download_history():
    return send_from_directory(EXPORT_DIR, EXPORT_FILENAME, as_attachment=True)


# Configuration
AUTO_SAVE_DIR = 'monitoring_data'  # Directory to save reports
AUTO_SAVE_INTERVAL = 300  # 5 minutes in seconds
MAX_RECENT_VISITS = 1000  #  to store more history
AUTO_SAVE_FILENAME = 'monitoring_report.xlsx'  # Single filename for auto-saves

# Data storage
file_access_data = {
    'recent_visits': [],
    'popular_files': defaultdict(int),
    'app_usage': defaultdict(int),
    'active_apps': defaultdict(int),
    'active_browsers': defaultdict(int),
    'last_save_time': None
}

# Common browser process names
BROWSERS = {
    'chrome': ['chrome', 'google-chrome', 'chrome.exe'],
    'firefox': ['firefox', 'firefox.exe'],
    'edge': ['msedge', 'msedge.exe'],
    'safari': ['safari', 'safari.exe'],
    'opera': ['opera', 'opera.exe'],
    'brave': ['brave', 'brave.exe']
}

def ensure_auto_save_dir():
    """Ensure the auto-save directory exists"""
    if not os.path.exists(AUTO_SAVE_DIR):
        os.makedirs(AUTO_SAVE_DIR)

def generate_excel_report():
    """Generate and save Excel report to file system """
    ensure_auto_save_dir()
    filepath = os.path.join(AUTO_SAVE_DIR, AUTO_SAVE_FILENAME)
    
    # Try to load existing workbook if it exists
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            # Remove all existing sheets
            for sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
        except Exception as e:
            print(f"Error loading existing workbook, creating new: {e}")
            wb = Workbook()
    else:
        wb = Workbook()
    
    # Sheet Active Applications
    ws1 = wb.create_sheet("Active Applications", 0)
    ws1.append(['Application', 'Process Count'])
    for app, count in file_access_data['active_apps'].items():
        ws1.append([app, count])
    
    # Sheet  Active Browsers
    ws2 = wb.create_sheet("Active Browsers", 1)
    ws2.append(['Browser', 'Process Count'])
    for browser, count in file_access_data['active_browsers'].items():
        ws2.append([browser, count])
    
    # Sheet  Recent Activity
    ws3 = wb.create_sheet("Recent Activity", 2)
    ws3.append(['Timestamp', 'File Name', 'File Path', 'Application'])
    for visit in file_access_data['recent_visits']:
        file_path = visit[0]
        file_name = os.path.basename(file_path)
        ws3.append([visit[1].strftime('%Y-%m-%d %H:%M:%S'), file_name, file_path, visit[2]])
    
    # Sheet  Popular Files
    ws4 = wb.create_sheet("Popular Files", 3)
    ws4.append(['File Name', 'File Path', 'Visit Count'])
    for file, count in sorted(file_access_data['popular_files'].items(), key=lambda x: x[1], reverse=True):
        file_name = os.path.basename(file)
        ws4.append([file_name, file, count])
    
    # Sheet  App Usage
    ws5 = wb.create_sheet("App Usage", 4)
    ws5.append(['Application', 'Files Opened'])
    for app, count in sorted(file_access_data['app_usage'].items(), key=lambda x: x[1], reverse=True):
        ws5.append([app, count])
    
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    wb.save(filepath)
    file_access_data['last_save_time'] = datetime.now()
    return AUTO_SAVE_FILENAME

def auto_save_worker():
    """Background worker for auto-saving reports (updates same file)"""
    while True:
        try:
            filename = generate_excel_report()
            print(f"Auto-saved monitoring report: {filename}")
        except Exception as e:
            print(f"Error auto-saving report: {e}")
        time.sleep(AUTO_SAVE_INTERVAL)

def detect_active_apps():
    """Detect currently running applications and browsers"""
    active_apps = defaultdict(int)
    active_browsers = defaultdict(int)
    
    for proc in psutil.process_iter(['name', 'exe', 'cmdline']):
        try:
            # Check for browsers
            is_browser = False
            for browser_name, browser_processes in BROWSERS.items():
                if proc.name().lower() in browser_processes or \
                   (proc.exe() and any(bp in proc.exe().lower() for bp in browser_processes)):
                    active_browsers[browser_name] += 1
                    is_browser = True
                    break
            
            # If not a browser, count as regular application
            if not is_browser:
                app_name = proc.name()
                active_apps[app_name] += 1
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    
    return active_apps, active_browsers

def get_system_files():
    """Get recently accessed files from the system"""
    system = platform.system()
    recent_files = []
    
    if system == 'Windows':
        # Windows recent files
        recent_dir = os.path.join(os.environ['APPDATA'], 'Microsoft', 'Windows', 'Recent')
        try:
            files = [os.path.join(recent_dir, f) for f in os.listdir(recent_dir)]
            files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            recent_files = files[:100]
        except Exception as e:
            print(f"Error accessing Windows recent files: {e}")
    
    elif system == 'Darwin':  # macOS
        # macOS recent files
        recent_dir = os.path.expanduser('~/Library/Application Support/com.apple.sharedfilelist')
        try:
            standard_locations = [
                '~/Library/Recent Documents',
                '~/Library/Containers/com.apple.finder/Data/Library/Recent Documents'
            ]
            for loc in standard_locations:
                path = os.path.expanduser(loc)
                if os.path.exists(path):
                    files = [os.path.join(path, f) for f in os.listdir(path)]
                    files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                    recent_files.extend(files[:50])
        except Exception as e:
            print(f"Error accessing macOS recent files: {e}")
    
    elif system == 'Linux':
        # Linux recent files (varies by desktop environment)
        locations = [
            '~/.local/share/recently-used.xbel',  # GNOME
            '~/.recently-used.xbel',             # Older GNOME
            '~/.local/share/RecentDocuments',     # KDE,
        ]
        try:
            for loc in locations:
                path = os.path.expanduser(loc)
                if os.path.exists(path):
                    if path.endswith('.xbel'):
                        with open(path, 'r') as f:
                            content = f.read()
                        files = [line.split('"')[1] for line in content.split('\n') 
                                if 'href="file://' in line]
                        recent_files.extend(files[:50])
                    else:
                        files = [os.path.join(path, f) for f in os.listdir(path)]
                        files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                        recent_files.extend(files[:50])
        except Exception as e:
            print(f"Error accessing Linux recent files: {e}")
    
    return recent_files

def perform_system_scan():
    """Immediately scan for recent files and active apps and update global state"""
    recent_files = get_system_files()
    now = datetime.now()
    
    # Update file access data
    for file_path in recent_files:
        file_access_data['recent_visits'].insert(0, (file_path, now, "System"))
        
        # Safely update popular files count
        if file_path not in file_access_data['popular_files']:
            file_access_data['popular_files'][file_path] = 0
        file_access_data['popular_files'][file_path] += 1
        
        if len(file_access_data['recent_visits']) > MAX_RECENT_VISITS:
            file_access_data['recent_visits'] = file_access_data['recent_visits'][:MAX_RECENT_VISITS]
    
    # Update application usage
    if 'System' not in file_access_data['app_usage']:
        file_access_data['app_usage']['System'] = 0
    file_access_data['app_usage']['System'] = len(recent_files)
    
    # Detect and update active applications and browsers
    active_apps, active_browsers = detect_active_apps()
    file_access_data['active_apps'] = active_apps
    file_access_data['active_browsers'] = active_browsers

    # Sync to Firebase if user is authenticated and in request context
    try:
        from flask import has_request_context
        if has_request_context():
            from flask_login import current_user
            if current_user.is_authenticated:
                app_summary = {
                    'active_apps': dict(active_apps),
                    'active_browsers': dict(active_browsers),
                    'recent_visits_count': len(recent_files)
                }
                FirebaseDB.save_application_usage(current_user.id, app_summary)
    except Exception as e:
        print(f"Firebase Sync Error (application usage): {e}")

def monitor_system():
    """Background thread to monitor system file access and active applications"""
    while True:
        try:
            perform_system_scan()
        except Exception as e:
            print(f"Error in monitor_system loop: {e}")
        time.sleep(60)  # Check every minute

# Start monitoring thread
monitor_thread = threading.Thread(target=monitor_system, daemon=True)
monitor_thread.start()

# Start auto-save thread
auto_save_thread = threading.Thread(target=auto_save_worker, daemon=True)
auto_save_thread.start()

@user_bp.route('/visited-applications', methods=['GET', 'POST'])
def visited_applications():
    # If Admin, fetch all users from Firebase
    all_users = []
    if current_user.is_authenticated and current_user.role == 'Admin':
        global_users_data = FirebaseDB.get_global_users()
        all_users = []
        for gu in global_users_data:
            if gu.get('role') == 'User' and gu.get('username'):
                all_users.append({
                    'username': gu['username'],
                    'email': gu.get('email', 'Global User')
                })
        all_users = sorted(all_users, key=lambda x: x['username'])
    
    if request.method == 'POST':
        try:
            perform_system_scan()
            flash('System scan completed successfully. Monitoring data updated.', 'success')
        except Exception as e:
            flash(f'Error performing system scan: {str(e)}', 'error')

    username = (
        request.form.get('username')
        or request.args.get('username')
        or (current_user.username if hasattr(current_user, "username") and current_user.is_authenticated else "Guest")
    )

    # Prepare data for template
    all_recent_visits = file_access_data['recent_visits']
    session_windows = get_user_session_windows(username)
    
    # Filter recent visits by session windows
    recent_visits = []
    filtered_popular_files = defaultdict(int)
    filtered_app_usage = defaultdict(int)
    
    offset = datetime.now() - datetime.utcnow()
    if session_windows:
        for visit in all_recent_visits:
            # visit format: (file_path, timestamp, application)
            visit_time = visit[1]
            # Compare local visit time against session windows converted to local time
            is_in_session = any((start + offset) <= visit_time <= (end + offset) for start, end in session_windows)
            
            if is_in_session:
                recent_visits.append(visit)
                filtered_popular_files[visit[0]] += 1
                filtered_app_usage[visit[2]] += 1
    else:
        recent_visits = all_recent_visits
        for visit in all_recent_visits:
            filtered_popular_files[visit[0]] += 1
            filtered_app_usage[visit[2]] += 1

    # Get top 10 popular files from filtered data
    popular_files = sorted(filtered_popular_files.items(), 
                         key=lambda x: x[1], reverse=True)[:10]
    
    # Get top 10 app usage from filtered data
    app_usage = sorted(filtered_app_usage.items(), 
                     key=lambda x: x[1], reverse=True)[:10]
    
    # Get active applications and browsers (These remain live/machine-wide as they reflect current state)
    active_apps = file_access_data['active_apps']
    active_browsers = file_access_data['active_browsers']
    
    return render_template('user/visited_applications.html',
                         recent_visits=recent_visits,
                         popular_files=popular_files,
                         app_usage=app_usage,
                         active_apps=active_apps,
                         active_browsers=active_browsers,
                         last_save_time=file_access_data['last_save_time'].strftime('%Y-%m-%d %H:%M:%S') if file_access_data['last_save_time'] else None,
                         auto_save_dir=AUTO_SAVE_DIR,
                         auto_save_filename=AUTO_SAVE_FILENAME,
                         username=username,
                         all_users=all_users)

@user_bp.route('/export-data')
def export_data():
    """Export all collected data to Excel (and update auto-save file)"""
    # Update the auto-save file
    generate_excel_report()
    
    # Then create response for download
    wb = Workbook()
    
    # Sheet 1 Active Applications
    ws1 = wb.active
    ws1.title = "Active Applications"
    ws1.append(['Application', 'Process Count'])
    for app, count in file_access_data['active_apps'].items():
        ws1.append([app, count])
    
    # Sheet 2 Active Browsers
    ws2 = wb.create_sheet(title="Active Browsers")
    ws2.append(['Browser', 'Process Count'])
    for browser, count in file_access_data['active_browsers'].items():
        ws2.append([browser, count])
    
    # Sheet 3 Recent Activity
    ws3 = wb.create_sheet(title="Recent Activity")
    ws3.append(['Timestamp', 'File Name', 'File Path', 'Application'])
    for visit in file_access_data['recent_visits']:
        file_path = visit[0]
        file_name = os.path.basename(file_path)
        ws3.append([visit[1].strftime('%Y-%m-%d %H:%M:%S'), file_name, file_path, visit[2]])
    
    # Sheet 4 Popular Files
    ws4 = wb.create_sheet(title="Popular Files")
    ws4.append(['File Name', 'File Path', 'Visit Count'])
    for file, count in sorted(file_access_data['popular_files'].items(), key=lambda x: x[1], reverse=True):
        file_name = os.path.basename(file)
        ws4.append([file_name, file, count])
    
    # Sheet 5 App Usage
    ws5 = wb.create_sheet(title="App Usage")
    ws5.append(['Application', 'Files Opened'])
    for app, count in sorted(file_access_data['app_usage'].items(), key=lambda x: x[1], reverse=True):
        ws5.append([app, count])
    
    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Create response
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=system_activity_report.xlsx'
    
    return response

@user_bp.route('/last-save-time')
def get_last_save_time():
    """Get the timestamp of the most recent auto-save"""
    try:
        filepath = os.path.join(AUTO_SAVE_DIR, AUTO_SAVE_FILENAME)
        if os.path.exists(filepath):
            last_save = datetime.fromtimestamp(os.path.getmtime(filepath))
            return jsonify({
                'last_save_time': last_save.strftime('%Y-%m-%d %H:%M:%S'),
                'filename': AUTO_SAVE_FILENAME
            })
        return jsonify({'last_save_time': None})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# @user_bp.template_filter('basename')
# def basename_filter(path):
#     return os.path.basename(path)

# Extract domain from URL
def extract_domain(url):
    match = re.search(r'(https?://)?(www\.)?([a-z0-9-]+)(\.[a-z]{2,}){1,2}', url)
    return match.group(3) if match else 'unknown'

# threat patterns
threat_patterns = {
    'betting': r'\b(bet|betting|gambl|casino|poker|sportsbook|wager|odds|slot)\b',
    'porn': r'\b(porn|xxx|adult|sex|fuck|nude|nsfw|webcam|escort)\b',
    'social': r'\b(facebook|instagram|twitter|tiktok|social[\s-]?media|linkedin)\b',
    'video': r'\b(youtube|vimeo|dailymotion|stream|twitch|netflix)\b',
    'trading': r'\b(forex|crypto|bitcoin|trading|binance|coinbase|kraken|mt[45])\b',
    'hacking': r'\b(hack|crack|keygen|cheat|exploit|bypass|ddos|injection)\b',
    'money': r'\b(western[\s-]?union|money[\s-]?gram|paypal|venmo|cash[\s-]?app)\b',
    'cloud': r'\b(dropbox|google[\s-]?drive|mega(?:\.nz)?|onedrive|box(?:\.com)?)\b',
    'anonymous': r'\b(tor|vpn|proxy|anonymous|incognito|hide[\s-]?ip)\b',
    'shopping': r'\b(amazon|ebay|alibaba|etsy|shopify|walmart|target)\b',
    'job': r'\b(linkedin|indeed|monster|career|job|employment)\b',
    'forum': r'\b(reddit|forum|4chan|discord|telegram|whatsapp)\b'
}

category_weights = {
    'betting': 0.9, 'porn': 1.0, 'trading': 0.8, 'hacking': 1.0,
    'money': 0.7, 'cloud': 0.6, 'anonymous': 0.8, 'social': 0.4,
    'video': 0.3, 'shopping': 0.2, 'job': 0.5, 'forum': 0.4
}

ml_df = None

def get_ml_df():
    global ml_df
    if ml_df is None:
        update_ml_pipeline()
    return ml_df

# Keep track of synced history IDs to avoid hitting Firebase quotas (429 errors)
synced_history_ids = set()

def update_ml_pipeline():
    global ml_df
    try:
        # Get all browser history from Firestore
        history_docs = FirebaseDB.get_all_browser_history(limit=5000)
        
        # Build user maps for username lookup and ID lookup
        global_users = FirebaseDB.get_global_users()
        user_map = {}
        uname_to_uid_map = {}
        for gu in global_users:
            uid = gu.get('id') or gu.get('user_id')
            uname = gu.get('username')
            if uid and uname:
                user_map[str(uid)] = uname
                uname_to_uid_map[uname] = uid
                
        # Parse into a list of dicts
        history_list = []
        for doc in history_docs:
            uid = doc.get('user_id')
            uname = doc.get('username') or doc.get('Username') or user_map.get(str(uid)) or 'Unknown'
            
            last_visit = doc.get('Last Visit') or doc.get('last_visit') or doc.get('last_visit_time')
            if isinstance(last_visit, datetime) and last_visit.tzinfo is not None:
                last_visit = last_visit.replace(tzinfo=None)
                
            history_list.append({
                'history_id': doc.get('history_id') or doc.get('id') or 'firestore_doc',
                'Username': uname,
                'Browser': doc.get('Browser') or doc.get('browser') or 'unknown',
                'URL': doc.get('URL') or doc.get('url') or '',
                'Title': doc.get('Title') or doc.get('title') or '',
                'Visit Count': doc.get('Visit Count') or doc.get('visit_count') or 1,
                'Last Visit': last_visit
            })
            
        import pandas as pd
        df = pd.DataFrame(history_list)
        
        if df.empty:
            ml_df = pd.DataFrame(columns=['history_id', 'Username', 'Browser', 'Title', 'URL', 'Visit Count', 'Last Visit', 'anomaly_score', 'is_anomaly', 'risk_level', 'composite_risk', 'threat_score'])
            for cat in threat_patterns.keys():
                ml_df[f'is_{cat}'] = 0
            return
            
        # Data cleaning
        df['Browser'] = df['Browser'].str.lower().str.strip().fillna('unknown')
        df['Title'] = df['Title'].str.lower().str.strip().fillna('')
        df['URL'] = df['URL'].str.lower().str.strip().fillna('')
        df['domain'] = df['URL'].apply(extract_domain)
        
        # Apply threat detection
        for category, pattern in threat_patterns.items():
            df[f'is_{category}'] = (
                df['Title'].str.contains(pattern, case=False, regex=True) | 
                df['URL'].str.contains(pattern, case=False, regex=True) |
                df['domain'].str.contains(pattern, case=False, regex=True)
            ).astype(int)
            
        df['Last Visit'] = pd.to_datetime(df['Last Visit'], errors='coerce')

        # Time-based features
        df['hour'] = df['Last Visit'].dt.hour
        df['weekday'] = df['Last Visit'].dt.weekday
        df['is_weekend'] = df['weekday'].isin([5, 6]).astype(int)
        df['late_night'] = ((df['hour'] >= 23) | (df['hour'] <= 5)).astype(int)

        # URL features
        df['url_length'] = df['URL'].str.len()
        df['url_depth'] = df['URL'].str.count('/') - 2
        df['has_query'] = df['URL'].str.contains(r'\?', regex=False).astype(int)

        from sklearn.ensemble import IsolationForest
        # Deterministic Browser Mapping
        BROWSER_MAP = {'chrome': 0, 'firefox': 1, 'edge': 2, 'safari': 3, 'opera': 4, 'unknown': 5}
        df['browser_code'] = df['Browser'].map(BROWSER_MAP).fillna(5).astype(int)
        df['visits_per_hour'] = df.groupby(
            [df['Last Visit'].dt.date, df['Last Visit'].dt.hour]
        )['Visit Count'].transform('sum')

        # Prepare features for Isolation Forest
        features = [
            'is_betting', 'is_porn', 'is_trading', 'is_hacking', 'is_money',
            'hour', 'is_weekend', 'late_night', 'url_length', 'url_depth',
            'has_query', 'visits_per_hour', 'browser_code'
        ]

        X = df[features]

        if not df.empty and len(df) >= 5:
            model = IsolationForest(
                n_estimators=150, max_samples=0.8, contamination='auto',
                random_state=42, n_jobs=-1
            )
            model.fit(X)
            df['anomaly_score'] = model.decision_function(X)
            # Corrected prediction: in IsolationForest, normal is 1 and anomaly is -1.
            # Convert to: Anomaly -> 1, Normal -> 0.
            predictions = model.predict(X)
            df['is_anomaly'] = (predictions == -1).astype(int)
        else:
            df['anomaly_score'] = 0.0
            df['is_anomaly'] = 0

        # Calculate threat score
        df['threat_score'] = 0
        for category in threat_patterns.keys():
            df['threat_score'] += df[f'is_{category}'] * category_weights.get(category, 0.5)

        # Combine with anomaly score
        df['composite_risk'] = 0.6 * df['threat_score'] + 0.4 * (1 - df['anomaly_score'])

        # Risk levels
        bins = [0, 0.3, 0.6, 0.8, 1]
        labels = ['Low', 'Medium', 'High', 'Critical']
        df['risk_level'] = pd.cut(df['composite_risk'], bins=bins, labels=labels)
        ml_df = df

        # Sync High/Critical anomalies to Firebase
        try:
            high_risk_df = df[df['risk_level'].isin(['High', 'Critical'])]
            if 'Last Visit' in high_risk_df.columns:
                high_risk_df = high_risk_df.sort_values(by='Last Visit', ascending=False)
                
            sync_count = 0
            for _, row in high_risk_df.iterrows():
                # Avoid redundant syncs to save quota
                h_id = row['history_id']
                if h_id in synced_history_ids:
                    continue
                    
                # Try to find the auth user id for this username from in-memory map
                uid = uname_to_uid_map.get(row['Username'], 0)
                
                # Limit to a maximum of 5 concurrent uploads per update to prevent Firestore quota exhaustion (429/504)
                if sync_count >= 5:
                    break
                    
                # Sync to Firebase in background
                import threading
                threading.Thread(
                    target=FirebaseDB.save_anomaly,
                    kwargs={
                        'user_id': uid,
                        'anomaly_type': "Behavioral Anomaly",
                        'severity': str(row['risk_level']),
                        'details': f"URL: {row['URL']} | Score: {row['composite_risk']:.2f}",
                        'doc_id': f"ml_{h_id}"
                    },
                    daemon=True
                ).start()
                synced_history_ids.add(h_id)
                sync_count += 1
        except Exception as e:
            print(f"Firebase Sync Error (ML anomalies): {e}")
    except Exception as e:
        print(f"Error updating ML pipeline: {e}")

@user_bp.route('/anomaly-detection', methods=['GET', 'POST'])
def anomaly_detection():
    df = get_ml_df()
    if current_user.role != 'Admin':
        df = df[df['Username'] == current_user.username]
    # Get the top 15 detected anomalies sorted by current date and time
    anomalies = df[['Username', 'Browser', 'Title', 'Last Visit', 'anomaly_score', 'is_anomaly']].sort_values('Last Visit', ascending=False).head(15)
    
    # Get the threat detection results
    threat_detection = df[[col for col in df.columns if col.startswith('is_')]].sum().sort_values(ascending=False)

    # Get high-risk activities sorted by current date and time
    high_risk = df[df['risk_level'].isin(['High', 'Critical'])][
        ['Username', 'Browser', 'Title', 'Last Visit', 'risk_level', 'composite_risk']
    ].sort_values('Last Visit', ascending=False).head(20)

    # === New: Behavior Analysis Summary ===
    total_records = len(df)
    total_anomalies = df['is_anomaly'].sum()
    anomaly_ratio = total_anomalies / total_records if total_records > 0 else 0

    if anomaly_ratio < 0.05:
        behavior_status = "Normal user behavior"
        behavior_message = "User browsing and system interactions appear typical with minimal risk detected."
        behavior_icon = "fa-check-circle"
        behavior_color = "green"
    elif 0.05 <= anomaly_ratio < 0.15:
        behavior_status = "Slightly unusual behavior"
        behavior_message = "Some unusual patterns were detected, but they may not indicate significant risk."
        behavior_icon = "fa-exclamation-circle"
        behavior_color = "orange"
    else:
        behavior_status = "Unusual or risky behavior patterns predicted"
        behavior_message = "Significant anomaly patterns were found indicating potential abnormal activity in the next 7 days."
        behavior_icon = "fa-radiation"
        behavior_color = "red"

    behavior_analysis = {
        "status": behavior_status,
        "message": behavior_message,
        "icon": behavior_icon,
        "color": behavior_color,
        "total_anomalies": int(total_anomalies),
        "total_records": int(total_records)
    }

    return render_template(
        'user/anomaly_detection.html',
        anomalies=anomalies.to_html(classes='table table-striped', index=False, border=0),
        threat_detection=threat_detection.to_frame(name='Count').to_html(classes='table table-striped', index=True, border=0),
        high_risk=high_risk.to_html(classes='table table-striped', index=False, border=0),
        behavior_analysis=behavior_analysis
    )



@user_bp.route('/reports-analytics', methods=['GET', 'POST'])
def reports_analytics():
    df = get_ml_df()
    if current_user.role != 'Admin':
        df = df[df['Username'] == current_user.username]

    

    # Prepare data for Chart.js
    risk_distribution = df['risk_level'].value_counts().to_dict()
    categories = [col for col in df.columns if col.startswith('is_')]
    avg_risk_by_category = df[categories].multiply(df['composite_risk'], axis=0).mean().sort_values(ascending=False).to_dict()

    return render_template(
        'user/reports_analytics.html',
        risk_distribution=risk_distribution,
        avg_risk_by_category=avg_risk_by_category
    )
    
@user_bp.route('/notifications')
def notifications():
    df = get_ml_df()
    if current_user.role != 'Admin':
        df = df[df['Username'] == current_user.username]
    target_categories = ['anonymous', 'betting', 'hacking', 'money', 'trading', 'porn']
    high_risk_df = df[
        (df['risk_level'].isin(['High', 'Critical'])) &
        (df[[f'is_{cat}' for cat in target_categories]].any(axis=1))
    ]
    
    notifications = []
    for _, row in high_risk_df.iterrows():
        triggered_categories = [
            cat for cat in target_categories 
            if row[f'is_{cat}'] == 1
        ]
        
        for category in triggered_categories:
            display_category = "Adult Content" if category == "porn" else category.capitalize()
            
            
            threat_score = row['threat_score']
            visit_count = row['Visit Count']

            if threat_score > 0.8:
                recommendation = (
                    f" Very high risk detected in {display_category}. "
                    f"Limit your time in this category to less than 1 hour per day and avoid frequent revisits."
                )
            elif 0.5 < threat_score <= 0.8:
                reduced_hours = max(1, round(visit_count * 0.5))
                recommendation = (
                    f" Elevated threat detected. Reduce your engagement in {display_category} "
                    f"by at least 50%. Limit to around {reduced_hours} hours weekly."
                )
            else:
                recommendation = (
                    f" Moderate risk level for {display_category}. Keep usage balanced and avoid excessive sessions."
                )

            notifications.append({
                'category': display_category,
                'title': row['Title'],
                'url': row['URL'],
                'risk': row['risk_level'],
                'timestamp': row['Last Visit'].strftime('%Y-%m-%d %H:%M:%S'),
                'visit_count': row['Visit Count'],
                'domain': row['domain'],
                'anomaly_score': round(row['anomaly_score'], 2),
                'threat_score': round(threat_score, 2),
                'is_sensitive': (category == 'porn'),
                'username': row.get('Username', 'Unknown'),
                'history_id': str(row.get('history_id', '')),
                'recommendation': recommendation  
            })
    
    
    # Sort strictly by timestamp descending (most recent first)
    notifications.sort(key=lambda x: -pd.to_datetime(x['timestamp']).timestamp())
    
    admin_messages = []
    read_history_ids = set()
    if current_user.is_authenticated:
        try:
            from firebase_config import db_firestore
            if db_firestore is not None:
                docs = db_firestore.collection('activities') \
                    .where('user_id', '==', current_user.id) \
                    .stream()
                
                activities = []
                for d in docs:
                    activities.append(d.to_dict())
                
                admin_acts = [act for act in activities if act.get('activity_type') == 'Admin Action']
                def get_ts(x):
                    t = x.get('timestamp')
                    if isinstance(t, datetime):
                        return t
                    return datetime.min
                admin_acts.sort(key=get_ts, reverse=True)
                
                admin_messages = []
                for act in admin_acts:
                    ts = act.get('timestamp')
                    ts_str = ts.strftime('%Y-%m-%d %H:%M:%S') if isinstance(ts, datetime) else str(ts)
                    admin_messages.append({'details': act.get('details'), 'timestamp': ts_str})
                    
                read_activity_type = 'Admin Alert Read' if current_user.role == 'Admin' else 'User Alert Read'
                read_activities = [act for act in activities if act.get('activity_type') == read_activity_type]
                read_history_ids = {str(act.get('details')) for act in read_activities}
        except Exception as e:
            print(f"Error fetching activities: {e}")
            
        notifications = [n for n in notifications if str(n['history_id']) not in read_history_ids]
    
    return render_template('user/notifications_alert.html', notifications=notifications, admin_messages=admin_messages)

@user_bp.context_processor
def inject_notification_count():
    df = get_ml_df()
    if current_user.is_authenticated and current_user.role != 'Admin':
        df = df[df['Username'] == current_user.username]
        
    target_categories = ['anonymous', 'betting', 'hacking', 'money', 'trading', 'porn']
    
    high_risk_df = df[
        (df['risk_level'].isin(['High', 'Critical'])) &
        (df[[f'is_{cat}' for cat in target_categories]].any(axis=1))
    ]
    
    if current_user.is_authenticated:
        try:
            from firebase_config import db_firestore
            if db_firestore is not None:
                read_activity_type = 'Admin Alert Read' if current_user.role == 'Admin' else 'User Alert Read'
                docs = db_firestore.collection('activities') \
                    .where('user_id', '==', current_user.id) \
                    .where('activity_type', '==', read_activity_type) \
                    .stream()
                read_history_ids = {str(d.to_dict().get('details')) for d in docs}
                if not high_risk_df.empty:
                    high_risk_df = high_risk_df[~high_risk_df['history_id'].astype(str).isin(read_history_ids)]
        except Exception as e:
            print(f"Error checking read alerts in context_processor: {e}")
            
    notification_count = len(high_risk_df)
    return dict(notification_count=notification_count)



def predict_user_behavior(file_path):
    # Load data
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path, parse_dates=['Last Visit'])
    else:
        df = pd.read_csv(file_path, encoding='latin1', parse_dates=['Last Visit'])

    # Clean data
    df['Browser'] = df['Browser'].str.lower().str.strip().fillna('unknown')
    df['Title'] = df['Title'].str.lower().str.strip().fillna('')
    df['URL'] = df['URL'].str.lower().str.strip().fillna('')

    # Extract domain
    def extract_domain(url):
        match = re.search(r'(https?://)?(www\.)?([a-z0-9-]+)(\.[a-z]{2,}){1,2}', url)
        return match.group(3) if match else 'unknown'

    df['domain'] = df['URL'].apply(extract_domain)

    # Threat patterns
    threat_patterns = {
        'betting': r'\b(bet|betting|gambl|casino|poker|sportsbook|wager|odds|slot)\b',
        'porn': r'\b(porn|xxx|adult|sex|fuck|nude|nsfw|webcam|escort)\b',
        'social': r'\b(facebook|instagram|twitter|tiktok|social[\s-]?media|linkedin)\b',
        'video': r'\b(youtube|vimeo|dailymotion|stream|twitch|netflix)\b',
        'trading': r'\b(forex|crypto|bitcoin|trading|binance|coinbase|kraken|mt[45])\b',
        'hacking': r'\b(hack|crack|keygen|cheat|exploit|bypass|ddos|injection)\b',
        'money': r'\b(western[\s-]?union|money[\s-]?gram|paypal|venmo|cash[\s-]?app)\b',
        'cloud': r'\b(dropbox|google[\s-]?drive|mega(?:\.nz)?|onedrive|box(?:\.com)?)\b',
        'anonymous': r'\b(tor|vpn|proxy|anonymous|incognito|hide[\s-]?ip)\b',
        'shopping': r'\b(amazon|ebay|alibaba|etsy|shopify|walmart|target)\b',
        'job': r'\b(linkedin|indeed|monster|career|job|employment)\b',
        'forum': r'\b(reddit|forum|4chan|discord|telegram|whatsapp)\b'
    }

    for category, pattern in threat_patterns.items():
        df[f'is_{category}'] = (
            df['Title'].str.contains(pattern, case=False, regex=True) |
            df['URL'].str.contains(pattern, case=False, regex=True) |
            df['domain'].str.contains(pattern, case=False, regex=True)
        ).astype(int)

    # Ensure Last Visit is datetime
    df['Last Visit'] = pd.to_datetime(df['Last Visit'], errors='coerce')
    # Drop invalid or missing dates
    df = df.dropna(subset=['Last Visit'])

    # Time-based features
    df['hour'] = df['Last Visit'].dt.hour
    df['weekday'] = df['Last Visit'].dt.weekday
    df['is_weekend'] = df['weekday'].isin([5, 6]).astype(int)
    df['late_night'] = ((df['hour'] >= 23) | (df['hour'] <= 5)).astype(int)

    # URL features
    df['url_length'] = df['URL'].str.len()
    df['url_depth'] = df['URL'].str.count('/') - 2
    df['has_query'] = df['URL'].str.contains(r'\?', regex=False).astype(int)

    # Browser features
    BROWSER_MAP = {'chrome': 0, 'firefox': 1, 'edge': 2, 'safari': 3, 'opera': 4, 'unknown': 5}
    df['browser_code'] = df['Browser'].map(BROWSER_MAP).fillna(5).astype(int)
    df['visits_per_hour'] = df.groupby(
        [df['Last Visit'].dt.date, df['Last Visit'].dt.hour]
    )['Visit Count'].transform('sum')

    # Features for model
    features = [
        'is_betting','is_porn','is_trading','is_hacking','is_money',
        'hour','is_weekend','late_night','url_length','url_depth',
        'has_query','visits_per_hour','browser_code'
    ]
    X = df[features]

    # Isolation Forest Anomaly Detection with robust sample guard
    if not df.empty and len(df) >= 5:
        model = IsolationForest(
            n_estimators=150, max_samples=0.8, contamination='auto',
            random_state=42, n_jobs=-1
        )
        model.fit(X)
        df['is_anomaly'] = model.predict(X)  # -1 = unusual, 1 = normal
    else:
        df['is_anomaly'] = 1  # Default to normal for small datasets

    # Map to readable behavior
    df['behavior'] = df['is_anomaly'].map({1: 'Normal', -1: 'Unusual'})

    # Return last record's behavior and features
    last_record = df.iloc[-1]
    behavior_status = last_record['behavior']
    current_features = last_record[features].to_dict()

    return behavior_status, current_features

@user_bp.route('/user-behaviour', methods=['GET', 'POST'])
def user_behavior():
    file_path = os.path.join(EXPORT_DIR, EXPORT_FILENAME)
    
    # Gracefully check if file exists to prevent crashing the server
    if not os.path.exists(file_path):
        flash('No telemetry records detected. Please perform an extraction scan first.', 'info')
        return render_template(
            'user/behaviour.html',
            behavior_status='No Data Available',
            features={f: 0 for f in [
                'is_betting','is_porn','is_trading','is_hacking','is_money',
                'hour','is_weekend','late_night','url_length','url_depth',
                'has_query','visits_per_hour','browser_code'
            ]}
        )
        
    try:
        behavior_status, features = predict_user_behavior(file_path)
    except Exception as e:
        flash(f'Telemetry Analysis Error: {str(e)}', 'error')
        return render_template(
            'user/behaviour.html',
            behavior_status='Error',
            features={}
        )
    
    return render_template(
        'user/behaviour.html',
        behavior_status=behavior_status,
        features=features
    )

@user_bp.route('/take-action', methods=['POST'])
@login_required
def take_action():
    data = request.get_json()
    username = data.get('username')
    action = data.get('action')

    if not username or not action:
        return jsonify({'status': 'error', 'message': 'Missing parameters'}), 400

    global_users = FirebaseDB.get_global_users()
    user_record = None
    for gu in global_users:
        if gu.get('username') == username:
            user_record = gu
            break
            
    if not user_record:
        return jsonify({'status': 'error', 'message': 'User not found'}), 404
        
    user_id = user_record.get('id') or user_record.get('user_id')

    if action == 'Warning':
        FirebaseDB.save_activity(
            user_id,
            'Admin Action',
            'Warning: High risk activity detected on your account. Please review your activity.'
        )
        return jsonify({'status': 'success', 'message': 'Warning sent.'})
        
    elif action == 'Suspend':
        FirebaseDB.save_activity(
            user_id,
            'Admin Action',
            'Account Suspended: Your account has been suspended due to severe policy violations.'
        )
        FirebaseDB.save_user(user_id, {'role': 'Suspended'})
        return jsonify({'status': 'success', 'message': 'User suspended.'})
        
    elif action == 'Resolve':
        FirebaseDB.save_activity(
            user_id,
            'Admin Action',
            'Threat Resolved: The previous security threat on your account has been marked as resolved.'
        )
        return jsonify({'status': 'success', 'message': 'Threat resolved.'})
        
    return jsonify({'status': 'error', 'message': 'Invalid action'}), 400

@user_bp.route('/mark-alert-read', methods=['POST'])
@login_required
def mark_alert_read():
    data = request.get_json()
    history_ids = data.get('history_ids', [])
    
    if not history_ids:
        return jsonify({'status': 'error', 'message': 'No history IDs provided'}), 400
        
    activity_type = 'Admin Alert Read' if current_user.role == 'Admin' else 'User Alert Read'
    
    try:
        from firebase_config import db_firestore
        if db_firestore is not None:
            docs = db_firestore.collection('activities') \
                .where('user_id', '==', current_user.id) \
                .where('activity_type', '==', activity_type) \
                .stream()
            existing_hids = {str(d.to_dict().get('details')) for d in docs}
            
            for hid in history_ids:
                if str(hid) not in existing_hids:
                    FirebaseDB.save_activity(current_user.id, activity_type, str(hid))
            
            return jsonify({'status': 'success', 'message': 'Alerts marked as read.'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
        
    return jsonify({'status': 'error', 'message': 'Database Connection Error'}), 500

# ==========================================
# SECURE FILE TRANSFER MODULE
# ==========================================
import os
from werkzeug.utils import secure_filename
from models import FileTransfer, User

SECURE_UPLOAD_FOLDER = os.path.join(os.getcwd(), 'secure_uploads')
os.makedirs(SECURE_UPLOAD_FOLDER, exist_ok=True)

def scan_file_for_threats(filename, filepath):
    """
    Heuristic threat detection engine for uploaded files.
    """
    dangerous_extensions = ['.exe', '.bat', '.vbs', '.ps1', '.js', '.cmd', '.scr', '.pif']
    ext = os.path.splitext(filename)[1].lower()
    
    if ext in dangerous_extensions:
        return True, f"Dangerous executable extension detected: {ext}"
        
    # Check file size (e.g. abnormally large scripts)
    size = os.path.getsize(filepath)
    if ext in ['.py', '.sh'] and size > 5 * 1024 * 1024:
        return True, "Script file is suspiciously large (>5MB)"
        
    return False, "Safe"

@user_bp.route('/secure-transfer', methods=['GET', 'POST'])
@login_required
def secure_transfer():
    if request.method == 'POST':
        receiver_email = request.form.get('receiver_email')
        file = request.files.get('file')
        
        if not receiver_email or not file or file.filename == '':
            flash('Recipient email and file are required.', 'error')
            log_data = {
                'user_id': current_user.id,
                'email': receiver_email,
                'action': 'file_transfer_failed',
                'ip_address': request.remote_addr,
                'user_agent': request.user_agent.string,
                'details': 'Missing recipient email or file.'
            }
            FirebaseDB.save_auth_log(log_data)
            return redirect(url_for('user.secure_transfer'))
            
        receiver = FirebaseDB.get_user_by_email(receiver_email)
        if not receiver:
            flash(f'No registered user found with email {receiver_email}. Transfers are strictly restricted to registered users.', 'error')
            log_data = {
                'user_id': current_user.id,
                'email': receiver_email,
                'action': 'file_transfer_failed',
                'ip_address': request.remote_addr,
                'user_agent': request.user_agent.string,
                'details': f'Unregistered recipient email: {receiver_email}'
            }
            FirebaseDB.save_auth_log(log_data)
            return redirect(url_for('user.secure_transfer'))
            
        filename = secure_filename(file.filename)
        safe_filename = f"{int(time.time())}_{filename}"
        file_path = os.path.join(SECURE_UPLOAD_FOLDER, safe_filename)
        file.save(file_path)
        
        # Threat Detection Scan
        is_threat, threat_details = scan_file_for_threats(filename, file_path)
        
        # Generate a unique Firestore ID for the transfer
        from firebase_config import db_firestore
        transfer_ref = db_firestore.collection('file_transfers').document()
        transfer_id = transfer_ref.id
        
        transfer_data = {
            'sender_id': current_user.id,
            'sender_username': current_user.username,
            'sender_email': current_user.email,
            'receiver_id': receiver.id,
            'receiver_username': receiver.username,
            'receiver_email': receiver.email,
            'filename': filename,
            'file_path': file_path,
            'file_size': os.path.getsize(file_path),
            'is_threat': is_threat,
            'threat_details': threat_details if is_threat else 'No threats detected',
            'status': 'blocked' if is_threat else 'delivered'
        }
        
        FirebaseDB.save_file_transfer(transfer_id, transfer_data)
        
        if is_threat:
            FirebaseDB.save_anomaly(
                user_id=current_user.id,
                anomaly_type="Suspicious File Transfer",
                severity="High",
                details=f"Transfer of {filename} was blocked due to threat detection."
            )
            log_data = {
                'user_id': current_user.id,
                'email': receiver_email,
                'action': 'file_transfer_threat_failed',
                'ip_address': request.remote_addr,
                'user_agent': request.user_agent.string,
                'details': f'Threat blocked: {threat_details} (File: {filename})'
            }
            FirebaseDB.save_auth_log(log_data)
            flash(f'Threat Intercepted: {threat_details}. The file has been quarantined.', 'error')
        else:
            flash('File successfully transferred and cleared security scans.', 'success')
            
        return redirect(url_for('user.secure_transfer'))

    # GET request - load inbox and outbox from Firestore
    inbox = []
    outbox = []
    from firebase_config import db_firestore
    if db_firestore is not None:
        try:
            inbox_docs = db_firestore.collection('file_transfers') \
                .where('receiver_id', '==', current_user.id) \
                .stream()
            outbox_docs = db_firestore.collection('file_transfers') \
                .where('sender_id', '==', current_user.id) \
                .stream()
            
            class Entity:
                def __init__(self, username, email):
                    self.username = username
                    self.email = email

            class TransferAdapter:
                def __init__(self, doc_id, d):
                    self.id = doc_id
                    self.sender_id = d.get('sender_id')
                    self.sender_username = d.get('sender_username') or d.get('sender_email') or 'Unknown'
                    self.receiver_id = d.get('receiver_id')
                    self.receiver_username = d.get('receiver_username') or d.get('receiver_email') or 'Unknown'
                    self.sender = Entity(d.get('sender_username') or d.get('sender_email') or 'Unknown', d.get('sender_email') or 'Unknown')
                    self.receiver = Entity(d.get('receiver_username') or d.get('receiver_email') or 'Unknown', d.get('receiver_email') or 'Unknown')
                    self.filename = d.get('filename', '')
                    self.file_path = d.get('file_path', '')
                    self.file_size = d.get('file_size', 0)
                    self.is_threat = d.get('is_threat', False)
                    self.threat_details = d.get('threat_details', '')
                    self.status = d.get('status', 'pending')
                    
                    t = d.get('timestamp')
                    if isinstance(t, datetime):
                        self.timestamp = t.replace(tzinfo=None) if t.tzinfo is not None else t
                    else:
                        self.timestamp = datetime.utcnow()
                        
            inbox = [TransferAdapter(d.id, d.to_dict()) for d in inbox_docs]
            outbox = [TransferAdapter(d.id, d.to_dict()) for d in outbox_docs]
            
            inbox.sort(key=lambda x: x.timestamp, reverse=True)
            outbox.sort(key=lambda x: x.timestamp, reverse=True)
        except Exception as e:
            print(f"Error loading transfers: {e}")
    
    return render_template('user/file_transfer.html', inbox=inbox, outbox=outbox)

@user_bp.route('/download-transfer/<string:transfer_id>')
@login_required
def download_transfer(transfer_id):
    from firebase_config import db_firestore
    if db_firestore is None:
        flash('Database connection failed.', 'error')
        return redirect(url_for('user.secure_transfer'))
        
    doc = db_firestore.collection('file_transfers').document(transfer_id).get()
    if not doc.exists:
        flash('Transfer record not found.', 'error')
        return redirect(url_for('user.secure_transfer'))
        
    transfer_data = doc.to_dict()
    
    # Security check: only receiver, sender or admin can download
    if current_user.id not in [transfer_data.get('receiver_id'), transfer_data.get('sender_id')] and current_user.role != 'Admin':
        flash('Unauthorized access.', 'error')
        return redirect(url_for('user.secure_transfer'))
        
    # Block downloading of threats unless it's an admin (or block entirely for safety)
    if transfer_data.get('is_threat') and current_user.role != 'Admin':
        flash('This file has been quarantined by the security system and cannot be downloaded.', 'error')
        return redirect(url_for('user.secure_transfer'))
        
    file_path = transfer_data.get('file_path')
    if not file_path or not os.path.exists(file_path):
        flash('File not found on server.', 'error')
        return redirect(url_for('user.secure_transfer'))
        
    return send_file(file_path, as_attachment=True, download_name=transfer_data.get('filename'))
